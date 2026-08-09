"""Dynamic Excel writer — appends transactions to any client template while
keeping the output identical in style to the input template.

Template preservation:
 - New rows copy every cell's style (font, borders, fill, number format)
   from the client's last existing data row.
 - Formula columns are replicated from the template's own formulas
   (row references translated), so NET / Balance UC / Check behave exactly
   like the client's existing rows.
 - Date cells match the template's type: text dates stay text in the same
   format; real dates keep the template's number format.
 - Year labels (SA / FY / ACC) are LEARNED from the client's existing rows
   (see core/year_labels.py) — every client's fiscal-year convention is
   reproduced, whatever it is.
 - The Analysis sheet is updated in place (values only) instead of being
   deleted and rebuilt, so its formatting survives.
"""

import re
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from core.financial_year import get_fy
from core.year_labels import YearLabeller, learn_year_columns


# ── Sheet / header helpers ────────────────────────────────────────────────────

def _detect_sheet(wb: openpyxl.Workbook, keyword: str) -> str:
    kw = keyword.lower()
    matches = [name for name in wb.sheetnames if kw in name.lower()]
    if not matches:
        raise KeyError(f"No sheet containing '{keyword}' found in {wb.sheetnames}")
    if kw == 'raw':
        # a sheet like 'Analysis 24 Raw (2)' is an analysis sheet, not the RAW tab
        pure = [m for m in matches if 'analysis' not in m.lower()]
        if pure:
            return pure[0]
    return matches[0]


def _read_headers(ws) -> dict:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {str(v).strip().lower(): i + 1 for i, v in enumerate(row1) if v is not None}


def _col(hdrs: dict, *names: str):
    for n in names:
        idx = hdrs.get(n.lower())
        if idx is not None:
            return idx
    return None


def _find_last_data_row(ws, key_cols: list) -> int:
    """Last row where any key column (date/description/no) has a value."""
    last = 1
    for r in range(2, ws.max_row + 1):
        for c in key_cols:
            if c and ws.cell(r, c).value not in (None, ''):
                last = r
                break
    return last


# ── Date handling ─────────────────────────────────────────────────────────────

_DATE_FMTS = ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %b %Y', '%d/%m/%y', '%m/%d/%Y')


def _detect_text_date_format(sample: str):
    s = str(sample).strip()
    for fmt in _DATE_FMTS:
        try:
            datetime.strptime(s, fmt)
            return fmt
        except ValueError:
            continue
    return None


# ── Formula translation ───────────────────────────────────────────────────────

_CELL_REF = re.compile(r'(\$?)([A-Z]{1,3})(\$?)(\d+)')


def _translate_formula(formula: str, src_row: int, dst_row: int) -> str:
    """Shift relative row references near src_row to dst_row.

    Absolute row refs ($4) and far-away refs (headers etc.) are kept."""
    delta = dst_row - src_row

    def repl(m):
        col_abs, col, row_abs, row_s = m.groups()
        row = int(row_s)
        if row_abs == '$':
            return m.group()
        if abs(row - src_row) <= 1:
            return f'{col_abs}{col}{row_abs}{row + delta}'
        return m.group()

    return _CELL_REF.sub(repl, formula)


# ── Client example extraction (teaches the AI this client's vocabulary) ──────

def extract_client_examples(template_path, sheet_name: str = None) -> list:
    """Return [{'description','type','reference','category'}] from the
    template's existing rows — the accountant's own past decisions."""
    wb = openpyxl.load_workbook(template_path, data_only=True)
    try:
        raw_name = sheet_name if (sheet_name and sheet_name in wb.sheetnames) \
            else _detect_sheet(wb, 'raw')
    except KeyError:
        return []
    ws = wb[raw_name]
    hdrs = _read_headers(ws)

    desc_col = _col(hdrs, 'Counter Party', 'Description', 'Transaction description',
                    'Details', 'Memo', 'Narrative')
    uc_col   = _col(hdrs, 'UC category', 'UC Category')
    ref_col  = _col(hdrs, 'Reference', 'Ref')
    type_col = _col(hdrs, 'Type', 'Transaction Type', 'Transaction type')

    if not desc_col or not uc_col:
        return []

    examples = []
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, desc_col).value
        cat  = ws.cell(r, uc_col).value
        if not desc or not cat:
            continue
        cat = str(cat).strip()
        if not cat or cat.startswith('='):
            continue
        examples.append({
            'description': str(desc).strip(),
            'type':      str(ws.cell(r, type_col).value or '').strip() if type_col else '',
            'reference': str(ws.cell(r, ref_col).value  or '').strip() if ref_col  else '',
            'category':  cat,
        })
    return examples


# ── Fallback year formatting (only when template has no existing rows) ───────

def _fallback_year_label(d, style: str):
    sa, acc = get_fy(d)          # sa='FY24' start-year, acc='24/25'
    end = acc[3:]
    if style == 'sa':
        return f"SA{acc}"        # 'SA24/25'
    if style == 'acc':
        return acc               # '24/25'
    if style == 'fy':
        return f"FY{end}"        # END-year label: FY25 = year 2024/25
    return None


# ── Main entry point ──────────────────────────────────────────────────────────

def write_workbook(
    transactions: list,
    categories: list,
    template_path,
    output_path,
    sheet_name: str = None,
) -> None:
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    # ── RAW sheet ─────────────────────────────────────────────────────────────
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[_detect_sheet(wb, 'raw')]

    hdrs = _read_headers(ws)

    no_col    = _col(hdrs, 'No', '#')
    sa_col    = _col(hdrs, 'SA')
    acc_col   = _col(hdrs, 'ACC', 'AC')
    fy_col    = _col(hdrs, 'FY')
    date_col  = _col(hdrs, 'Date')
    desc_col  = _col(hdrs, 'Counter Party', 'Description', 'Transaction description',
                     'Details', 'Memo', 'Narrative')
    ref_col   = _col(hdrs, 'Reference', 'Ref')
    type_col  = _col(hdrs, 'Type', 'Transaction Type', 'Transaction type')
    in_col    = _col(hdrs, 'Paid in', 'Paid In', 'IN', 'In', 'In (£)', 'In (GBP)')
    out_col   = _col(hdrs, 'Paid out', 'Paid Out', 'OUT', 'Out', 'Out (£)', 'Out (GBP)')
    amt_col   = _col(hdrs, 'Amount (GBP)', 'Amount (£)', 'Amount')
    bal_col   = _col(hdrs, 'Balance', 'Balance (GBP)', 'Balance (£)')
    net_col   = _col(hdrs, 'NET', 'Net')
    buc_col   = _col(hdrs, 'Balance UC')
    chk_col   = _col(hdrs, 'Check UC', 'Check')
    csv_cat_col = _col(hdrs, 'Spending Category', 'Category name')
    uc_col    = _col(hdrs, 'UC category', 'UC Category')
    ts_col     = _col(hdrs, 'Timestamp')
    from_col   = _col(hdrs, 'From')
    to_col     = _col(hdrs, 'To')
    status_col = _col(hdrs, 'Status')
    tag_col    = _col(hdrs, 'Tag 1', 'Tag')

    last_row = _find_last_data_row(ws, [date_col, desc_col, no_col])

    # ── Learn this client's year-label conventions from their own rows ───────
    labellers = learn_year_columns(
        ws, {'sa': sa_col, 'acc': acc_col, 'fy': fy_col}, date_col, last_row,
    )

    # ── Template rows used as style/formula models for appended rows ─────────
    model_row = last_row if last_row > 1 else None
    max_col = ws.max_column

    model_styles   = {}
    model_formulas = {}   # col -> (formula, source_row)
    date_text_fmt  = None
    if model_row:
        for c in range(1, max_col + 1):
            cell = ws.cell(model_row, c)
            model_styles[c] = copy(cell._style)
        # Collect the most recent formula per column, scanning several rows up:
        # some templates carry a formula only on the row-sign it applies to
        # (e.g. IN = IF(H>0,H,"") appears only on income rows).
        scan_from = max(2, model_row - 200)
        for r_scan in range(model_row, scan_from - 1, -1):
            for c in range(1, max_col + 1):
                if c in model_formulas:
                    continue
                v = ws.cell(r_scan, c).value
                if isinstance(v, str) and v.startswith('='):
                    model_formulas[c] = (v, r_scan)
        d_model = ws.cell(model_row, date_col).value if date_col else None
        if isinstance(d_model, str):
            date_text_fmt = _detect_text_date_format(d_model)

    # Sign-conditional IN/OUT formulas: when BOTH sides carry a formula in the
    # template, each appended row gets only the side matching its sign — the
    # other cell stays empty, exactly like the client's own rows.
    sign_conditional = (
        in_col in model_formulas and out_col in model_formulas
        and amt_col is not None
    )

    # ── Highest existing sequence number ──────────────────────────────────────
    last_no = 0
    if no_col:
        for r in range(2, last_row + 1):
            v = ws.cell(r, no_col).value
            try:
                last_no = max(last_no, int(v))
            except (TypeError, ValueError):
                pass

    def year_label(kind: str, fallback_style: str, d):
        lab: YearLabeller = labellers.get(kind)
        if lab:
            return lab.label_for(d)
        return _fallback_year_label(d, fallback_style)

    # ── Write transaction rows ────────────────────────────────────────────────
    for i, (txn, cat) in enumerate(zip(transactions, categories)):
        r = last_row + 1 + i
        seq = last_no + 1 + i

        d = txn['date']
        if isinstance(d, datetime):
            d = d.date()

        money_in  = txn.get('money_in', 0.0)
        money_out = txn.get('money_out', 0.0)
        balance   = txn.get('balance')
        desc      = (txn.get('description') or '').strip() or (txn.get('subcategory') or '').strip()

        # 1. apply the template's styles to the whole new row first
        for c in range(1, max_col + 1):
            if c in model_styles:
                ws.cell(r, c)._style = copy(model_styles[c])

        def w(col, val):
            if col and val is not None and col not in model_formulas:
                ws.cell(r, col, value=val)

        # 2. data values
        w(no_col, seq)
        if sa_col:
            w(sa_col, year_label('sa', 'sa', d))
        if acc_col:
            w(acc_col, year_label('acc', 'acc', d))
        if fy_col:
            w(fy_col, year_label('fy', 'fy', d))

        if date_col and date_col not in model_formulas:
            if date_text_fmt:
                ws.cell(r, date_col, value=datetime(d.year, d.month, d.day).strftime(date_text_fmt))
            else:
                ws.cell(r, date_col, value=datetime(d.year, d.month, d.day))

        w(desc_col, desc or None)
        w(ref_col,  (txn.get('reference') or '').strip() or None)
        w(type_col, (txn.get('subcategory') or '').strip() or None)
        w(in_col,   money_in  if money_in  > 0 else None)
        w(out_col,  money_out if money_out > 0 else None)
        w(bal_col,  balance)
        w(csv_cat_col, (txn.get('csv_category') or '').strip() or None)
        w(uc_col,   cat)
        w(ts_col,     txn.get('timestamp'))
        w(from_col,   (txn.get('from')   or '').strip() or None)
        w(to_col,     (txn.get('to')     or '').strip() or None)
        w(status_col, (txn.get('status') or '').strip() or None)
        w(tag_col,    (txn.get('tag')    or '').strip() or None)

        # signed amount column
        if amt_col and amt_col not in model_formulas:
            signed = money_in if money_in > 0 else (-money_out if money_out > 0 else None)
            if signed is not None:
                ws.cell(r, amt_col, value=signed)

        # 3. replicate the template's own formulas (NET, Balance UC, Check, …)
        for c, (f, src_row) in model_formulas.items():
            if sign_conditional:
                if c == in_col and money_in <= 0:
                    continue    # leave IN empty on money-out rows
                if c == out_col and money_out <= 0:
                    continue    # leave OUT empty on money-in rows
            ws.cell(r, c, value=_translate_formula(f, src_row, r))

        # 4. fallback formulas when the template rows carry static values
        if net_col and net_col not in model_formulas and in_col and out_col:
            in_l, out_l = get_column_letter(in_col), get_column_letter(out_col)
            ws.cell(r, net_col, value=f'={in_l}{r}-{out_l}{r}')
        if buc_col and buc_col not in model_formulas and net_col:
            buc_l, net_l = get_column_letter(buc_col), get_column_letter(net_col)
            if r - 1 >= 2:
                ws.cell(r, buc_col, value=f'={buc_l}{r - 1}+{net_l}{r}')
            else:
                ws.cell(r, buc_col, value=f'={net_l}{r}')
        if chk_col and chk_col not in model_formulas and bal_col and buc_col and balance is not None:
            bal_l, buc_l = get_column_letter(bal_col), get_column_letter(buc_col)
            ws.cell(r, chk_col, value=f'={bal_l}{r}-{buc_l}{r}')

    new_last = last_row + len(transactions)

    # ── Update Analysis sheet in place ────────────────────────────────────────
    try:
        analysis_name = _detect_sheet(wb, 'analysis')
    except KeyError:
        analysis_name = None

    pivot_year_col = fy_col or acc_col or sa_col
    net_source_col = net_col or amt_col

    if analysis_name and pivot_year_col and net_source_col and uc_col:
        _update_analysis(wb[analysis_name], ws, pivot_year_col, net_source_col,
                         uc_col, last_data_row=new_last)
        print(f"  [excel_writer] Updated {analysis_name} in place.")
    elif analysis_name:
        print("  [excel_writer] Missing year/net/category columns — Analysis left untouched.")

    wb.save(output_path)
    print(f"  [excel_writer] Saved -> {output_path}")


# ── Analysis update (values only — formatting preserved) ─────────────────────

def _update_analysis(ws_a, raw_ws, year_col, net_col, uc_col, last_data_row) -> None:
    raw_name    = raw_ws.title
    year_letter = get_column_letter(year_col)
    net_letter  = get_column_letter(net_col)
    uc_letter   = get_column_letter(uc_col)

    years_seen, cats = {}, []
    for r in range(2, last_data_row + 1):
        y = raw_ws.cell(r, year_col).value
        c = raw_ws.cell(r, uc_col).value
        if y is not None and str(y).strip() and not str(y).startswith('='):
            years_seen.setdefault(str(y).strip(), y)   # keep original type (str/int)
        if c is not None and str(c).strip() and not str(c).startswith('='):
            cs = str(c).strip()
            if cs not in cats:
                cats.append(cs)
    years = [years_seen[k] for k in sorted(years_seen)]
    cats.sort(key=str.lower)

    old_max_row = ws_a.max_row
    old_max_col = ws_a.max_column

    # style models taken from the existing populated area
    hdr_style = copy(ws_a.cell(4, 2)._style)
    cat_style = copy(ws_a.cell(5, 1)._style)
    val_style = copy(ws_a.cell(5, 2)._style)

    # clear old values but keep every cell's formatting
    for r in range(1, old_max_row + 1):
        for c in range(1, old_max_col + 1):
            ws_a.cell(r, c).value = None

    ws_a.cell(3, 1, 'Sum of NET')
    ws_a.cell(3, 2, 'Column Labels')
    ws_a.cell(4, 1, 'Row Labels')
    for j, yr in enumerate(years):
        cell = ws_a.cell(4, 2 + j)
        cell.value = yr
        cell._style = copy(hdr_style)

    for k, cat in enumerate(cats):
        r = 5 + k
        cell = ws_a.cell(r, 1)
        cell.value = cat
        cell._style = copy(cat_style)
        for j, yr in enumerate(years):
            yr_ref = f"${get_column_letter(2 + j)}$4"
            formula = (
                f"=SUMIFS('{raw_name}'!${net_letter}:${net_letter},"
                f"'{raw_name}'!${year_letter}:${year_letter},{yr_ref},"
                f"'{raw_name}'!${uc_letter}:${uc_letter},$A{r})"
            )
            cell = ws_a.cell(r, 2 + j)
            cell.value = formula
            cell._style = copy(val_style)

    total_r = 5 + len(cats)
    cell = ws_a.cell(total_r, 1)
    cell.value = 'Grand Total'
    cell._style = copy(cat_style)
    for j in range(len(years)):
        col_letter = get_column_letter(2 + j)
        cell = ws_a.cell(total_r, 2 + j)
        cell.value = f"=SUM({col_letter}5:{col_letter}{total_r - 1})"
        cell._style = copy(val_style)
