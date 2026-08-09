"""Build core/learned_examples.json from client-corrected feedback workbooks.

The client reviews output files and writes the CORRECT category in the column
to the right of the AI's 'UC Category' column. This script extracts those
corrections as (payee-pattern -> category) examples the AI learns from.

Run:  python scripts/learn_from_feedback.py [Feedbacks-folder]
"""

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))


def _payee_key(desc: str) -> str:
    """Normalise a description to a stable payee key: drop dates, refs, digits."""
    s = str(desc).upper()
    s = re.sub(r'\d+', ' ', s)                     # numbers out
    s = re.sub(r'[^A-Z& ]+', ' ', s)               # punctuation out
    words = [w for w in s.split() if len(w) > 1]
    # Drop common transaction-mechanics words that aren't the payee
    STOP = {'CARD', 'PAYMENT', 'TO', 'ON', 'DIRECT', 'DEBIT', 'CREDIT', 'AUTOMATED',
            'ONLINE', 'TRANSACTION', 'VIA', 'MOBILE', 'FP', 'THE', 'REF', 'BGC',
            'TFR', 'DD', 'SO', 'FASTER', 'INWARD', 'OUTWARD', 'BACS', 'RECEIVED',
            'TRANSFER', 'PYMT', 'IN', 'OUT', 'LVP'}
    words = [w for w in words if w not in STOP]
    return ' '.join(words[:4])


def _find_uc_col(ws) -> int | None:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and 'uc categ' in str(v).strip().lower():
            return c
    return None


def learn(folder: Path) -> dict:
    votes: dict[str, Counter] = defaultdict(Counter)
    sample_desc: dict[str, str] = {}

    for path in sorted(folder.glob('*.xlsx')):
        wb = openpyxl.load_workbook(path)
        for sheet in wb.sheetnames:
            if 'raw' not in sheet.lower():
                continue
            ws = wb[sheet]
            uc_col = _find_uc_col(ws)
            if not uc_col:
                continue
            # description column: first header matching known names
            desc_col = None
            for c in range(1, ws.max_column + 1):
                v = str(ws.cell(1, c).value or '').strip().lower()
                if v in ('counter party', 'description', 'details', 'transaction description',
                         'memo', 'narrative'):
                    desc_col = c
                    break
            if not desc_col:
                continue

            for r in range(2, ws.max_row + 1):
                desc = ws.cell(r, desc_col).value
                if not desc or not str(desc).strip():
                    continue
                # correction = first non-empty cell right of UC col
                corr = None
                for c in range(uc_col + 1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip():
                        corr = str(v).strip()
                        break
                if not corr or corr.startswith('='):   # skip formulas / junk
                    continue
                key = _payee_key(desc)
                if not key:
                    continue
                votes[key][corr] += 1
                sample_desc.setdefault(key, str(desc)[:70])

    # Majority vote per payee; drop ambiguous ties
    examples = {}
    for key, counter in votes.items():
        (top_cat, top_n), *rest = counter.most_common()
        total = sum(counter.values())
        if top_n / total >= 0.6:   # clear majority only
            examples[key] = {'category': top_cat, 'seen': total,
                             'example': sample_desc[key]}

    return examples


if __name__ == '__main__':
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / 'Feedbacks'
    examples = learn(folder)
    out = ROOT / 'core' / 'learned_examples.json'
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(examples, f, indent=1, ensure_ascii=False)
    print(f'Learned {len(examples)} payee patterns -> {out}')
    for k, v in sorted(examples.items(), key=lambda kv: -kv[1]['seen'])[:25]:
        print(f"  {v['seen']:3}x {k[:35]:<37} -> {v['category']}")
